import os.path
import openpyxl as ox
import sqlite3 as sql
import tkinter as tk
import tkinter.messagebox as mb
import threading as trg
import http.client
import subprocess as sbp


class DatabaseSQL:
    def __init__(self, name):
        self.conn = sql.connect(name)
        self.cursor = self.conn.cursor()

    def createtable(self, name, fields):
        self.cursor.execute('CREATE TABLE IF NOT EXISTS ' + name + ' (' + fields + ')')
        self.conn.commit()

    @staticmethod
    def insert_zaprose(fields, name):
        zp1 = 'INSERT INTO ' + name + ' ('
        zp2 = ') VALUES ('
        for d in range(0, len(fields) - 1):
            zp1 = zp1 + fields[d] + ', '
            zp2 = zp2 + '?, '
        zp1 = zp1 + fields[len(fields) - 1]
        zp2 = zp2 + '?)'
        return zp1 + zp2

    def insert(self, zaproce, value):
        self.cursor.execute(zaproce, value)
        self.conn.commit()

    def insert_many(self, zaproce, values):
        for d in values:
            self.cursor.execute(zaproce, d)
            self.conn.commit()

    def update(self, name, fields, yslovies, values):
        self.cursor.execute('UPDATE ' + name + ' SET ' + update_part(fields) + ' WHERE '
                            + update_part(yslovies), values)
        self.conn.commit()

    def delete(self, name, yslovies, values):
        self.cursor.execute('DELETE FROM ' + name + ' WHERE ' + update_part(yslovies), values)
        self.conn.commit()

    def select(self, stolbiki, table):
        self.cursor.execute('SELECT ' + stolbiki + ' FROM ' + table)
        return self.cursor.fetchall()

    def select_all(self, table):
        return self.select('*', table)

    def select_where(self, stolbiki, table, yslovies, values):
        self.cursor.execute('SELECT ' + stolbiki + ' FROM ' + table + ' WHERE ' + update_part(yslovies), values)
        return self.cursor.fetchall()

    def select_where_count(self, stolbiki, table, yslovies, values):
        if yslovies == "":
            self.cursor.execute('SELECT COUNT(' + stolbiki + ') FROM ' + table)
        else:
            self.cursor.execute('SELECT COUNT(' + stolbiki + ') FROM ' + table +
                                ' WHERE ' + update_part(yslovies), values)
        return self.cursor.fetchone()

    def select_all_where(self, table, yslovies, values):
        return self.select_where('*', table, yslovies, values)

    def close(self):
        self.cursor.close()
        self.conn.close()


def update_part(dfs):
    try:
        return dfs + ' = ?'
    except Exception:
        zp = ''
        for d in range(0, len(dfs) - 1):
            zp = zp + dfs[d] + ' = ? AND '
        zp = zp + dfs[len(dfs) - 1] + ' = ?'
        return zp


class DatabaseSoli(DatabaseSQL):
    def __init__(self):
        super().__init__('Soli.db')
        self.createtable("Soli", "id INTEGER PRIMARY KEY AUTOINCREMENT, solt TEXT NOT NULL")

    def insert_solt(self, value):
        self.insert(self.insert_zaprose('solt', 'Soli'), value)

    def insert_solt_many(self, values):
        self.insert_many(self.insert_zaprose('solt', 'Soli'), values)


class DatabasePaper(DatabaseSQL):
    def __init__(self):
        super().__init__('Paper.db')
        self.createtable("Paper", "id INTEGER PRIMARY KEY AUTOINCREMENT, paper TEXT NOT NULL")

    def insert_paper(self, value):
        self.insert(self.insert_zaprose('paper', 'Paper'), value)

    def insert_paper_many(self, values):
        self.insert_many(self.insert_zaprose('paper', 'Paper'), values)


class DatabaseSaltPaper(DatabaseSQL):
    def __init__(self):
        super().__init__('SaltPaper.db')
        self.createtable("SaltPaper", "id INTEGER PRIMARY KEY AUTOINCREMENT, saltpaperpaper TEXT NOT NULL")

    def insert_salt_paper(self, value):
        self.insert(self.insert_zaprose('saltpaper', 'SaltPaper'), value)

    def insert_salt_paper_many(self, values):
        self.insert_many(self.insert_zaprose('saltpaper', 'SaltPaper'), values)


class DatabaseUsers(DatabaseSQL):
    def __init__(self):
        super().__init__("Users.db")
        self.createtable("Users", "id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, hache TEXT NOT NULL")


class DatabaseMain(DatabaseSQL):
    def __init__(self):
        super().__init__("Budilniki.db")
        self.createtable("Budilniki", "id INTEGER PRIMARY KEY AUTOINCREMENT, Userid INTEGER NOT NULL, " +
                         "name TEXT NOT NULL, datetime TEXT NOT NULL, povtors INTEGER NOT NULL")


class DatabaseRaspisanieRead(DatabaseSQL):
    def __init__(self):
        super().__init__('Raspisanie.db')
        self.createtable("DayOfWeek", "id INTEGER PRIMARY KEY AUTOINCREMENT, day TEXT NOT NULL, dayseek TEXT NOT NULL")
        self.createtable("PairsTime", "id INTEGER PRIMARY KEY, time TEXT NOT NULL")
        self.createtable("Raspisanie", "id INTEGER PRIMARY KEY AUTOINCREMENT, id_pari INTEGER NOT NULL, " +
                         "day TEXT NOT NULL, para TEXT NOT NULL, cz TEXT NOT NULL, gr TEXT NOT NULL")
        self.createtable("GroupsOnN", "id INTEGER PRIMARY KEY AUTOINCREMENT, gr TEXT NOT NULL, " +
                                      "np TEXT NOT NULL, crs TEXT NOT NULL")
        self.createtable("ChislitelZnamenatelWeeks", "id INTEGER PRIMARY KEY AUTOINCREMENT, " +
                         "week TEXT NOT NULL, cz TEXT NOT NULL")
        self.insert_into_days = self.insert_zaprose(['day', 'dayseek'], 'DayOfWeek')
        self.insert_into_time = self.insert_zaprose(['id', 'time'], 'PairsTime')
        self.insert_into_pari = self.insert_zaprose(['id_pari', 'day', 'para', 'cz', 'gr'], 'Raspisanie')
        self.insert_into_groups = self.insert_zaprose(['gr', 'np', 'crs'], 'GroupsOnN')
        self.insert_into_weeks = self.insert_zaprose(['week', 'cz'], 'ChislitelZnamenatelWeeks')


class DatabaseRaspisanie(DatabaseRaspisanieRead):
    def __init__(self, days, dayss):
        super().__init__()
        for d in range(0, min(len(days), len(dayss))):
            if self.select_where_count('*', 'DayOfWeek', 'day', [days[d]]) == (0,):
                self.insert(self.insert_into_days, (days[d], dayss[d]))


class ExcelFunctionsDefault:
    def __init__(self):
        self.days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
        self.dayss = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб"]

    def ltr(self, tre):
        for d in range(0, min(len(self.days), len(self.dayss))):
            if self.days[d] == tre:
                return self.dayss[d]
            if self.dayss[d] == tre:
                return self.days[d]
        return ""

    @staticmethod
    def cell_merged_diapasone(mr, row, col):
        for d in mr:
            if d.min_row <= row <= d.max_row and d.min_col <= col <= d.max_col:
                return [d.min_row, d.min_col, d.max_row, d.max_col]
        return [row, col, row, col]

    @staticmethod
    def lishnie_probeli(st, t):
        try:
            hj = st.split(t)
            gf = hj[0].strip(" ")
            for d in range(1, len(hj)):
                gf = (gf + t + hj[d].strip(" ")).strip(" ")
            return gf
        except Exception:
            try:
                return st.strip(" ")
            except Exception:
                return st

    @staticmethod
    def sell_empty(cl):
        return cl.value is None

    def fact_max_size(self, sh):
        min_r = 0 - 1
        min_c = 0 - 1
        max_r = 1
        max_c = 1
        for d in range(1, sh.max_row + 1):
            for t in range(1, sh.max_column + 1):
                if not self.sell_empty(sh.cell(row=d, column=t)):
                    if d <= min_r or min_r == -1:
                        min_r = d
                    if t <= min_c or min_c == -1:
                        min_c = t
                    if d >= max_r:
                        max_r = d
                    if t >= max_c:
                        max_c = t
        return [min_r, min_c, max_r, max_c]

    def koordinates(self, sh, tx, f, ms):
        for d in range(ms[0], ms[2] + 1):
            for t in range(ms[1], ms[3] + 1):
                if not self.sell_empty(sh.cell(row=d, column=t)):
                    if self.lishnie_probeli(str(sh.cell(row=d, column=t).value), f) == tx:
                        return [d, t]
        return [-1, -1]

    def napravlenya(self, sh, ms, mr):
        dfr = []
        rt = self.koordinates(sh, "Направление подготовки", " ", ms)
        if rt[0] == -1:
            return []
        for t in range(max(rt[1] + 1, ms[1]), ms[3] + 1):
            dfrg = []
            if not self.sell_empty(sh.cell(row=rt[0], column=t)):
                dfrg.append(sh.cell(row=rt[0], column=t).value)
                dfrg.append(self.cell_merged_diapasone(mr, rt[0], t))
                dfr.append(dfrg)
        return dfr

    def groups_recursion_body(self, sh, ms, mr, rt, dfre):
        dfr = []
        dfr.extend(dfre)
        yt = self.cell_merged_diapasone(mr, rt[0], rt[1])
        dfrt = []
        for d in dfr:
            for t in range(max(ms[0], yt[0]), min(ms[2] + 1, yt[2] + 1)):
                for k in range(max(ms[1], d[1][1]), min(ms[3] + 1, d[1][3] + 1)):
                    dfrg = []
                    if not self.sell_empty(sh.cell(row=t, column=k)):
                        dfrg.append(sh.cell(row=t, column=k).value)
                        dfrg.append(self.cell_merged_diapasone(mr, t, k))
                        dfrt.append(dfrg)
        if not dfrt:
            rt[0] = rt[0] - 1
            rt[2] = rt[2] - 1
            if rt[0] == -1:
                return []
            return self.groups_recursion_body(sh, ms, mr, rt, dfre)
        dfr.extend(dfrt)
        return dfr

    def groups(self, sh, ms, mr):
        np = self.napravlenya(sh, ms, mr)
        dfr = []
        dfr.extend(np)
        kjh = [dfr[0][1][0], dfr[0][1][2]]
        rt = self.koordinates(sh, "Направленность/специализация", "/", ms)
        if rt[0] == -1:
            return dfr
        yt = self.cell_merged_diapasone(mr, rt[0], rt[1])
        if yt[0] - np[0][1][2] > 1:
            yt[0] = np[0][1][2] + 1
        dfr = self.groups_recursion_body(sh, ms, mr, yt, np)
        if not dfr:
            return []
        for d in range(0, len(dfr)):
            if kjh[0] <= dfr[d][1][0] <= kjh[1] or kjh[0] <= dfr[d][1][2] <= kjh[1]:
                dfr[d].append([""])
            else:
                fg = []
                for t in range(0, len(dfr)):
                    if dfr[d][1][0] - dfr[t][1][0] == 1 or dfr[d][1][2] - dfr[t][1][2] == 1:
                        if dfr[d][1][1] >= dfr[t][1][1] and dfr[d][1][3] <= dfr[t][1][3]:
                            fg.append(dfr[t][0])
                if len(fg) == 0:
                    if not (kjh[0] <= dfr[d][1][0] - 1 <= kjh[1] or kjh[0] <= dfr[d][1][2] - 1 <= kjh[1]):
                        for t in range(0, len(dfr)):
                            if dfr[d][1][0] - dfr[t][1][0] == 2 or dfr[d][1][2] - dfr[t][1][2] == 2:
                                if dfr[d][1][1] >= dfr[t][1][1] and dfr[d][1][3] <= dfr[t][1][3]:
                                    for k in dfr:
                                        if len(k) >= 3:
                                            for u in k[2]:
                                                if u == dfr[t][0]:
                                                    fg.append(k[0])
                                    dfr[d].append(fg)
                                    dfr[d].append(dfr[t][0])
                        if len(fg) == 0:
                            dfr.append([""])
                    else:
                        dfr[d].append([""])
                else:
                    dfr[d].append(fg)
        dfrt = []
        for d in range(0, len(dfr)):
            if len(dfr[d]) >= 3:
                if len(dfr[d][2]) > 1:
                    for t in dfr[d][2]:
                        for k in dfr:
                            if t == k[0]:
                                dfrt.append([str(dfr[d][0]) + " (" + t + ")",
                                             [dfr[d][1][0], max(dfr[d][1][1], k[1][1]),
                                              dfr[d][1][2], min(dfr[d][1][3], k[1][3])],
                                             str(dfr[d][0])])
                                break
                    dfr[d].remove(dfr[d][2])
                else:
                    yu = dfr[d][2][0]
                    dfr[d].remove(dfr[d][2])
                    dfr[d].append(yu)
        dfr.extend(dfrt)
        dfrt = []
        for d in dfr:
            dfrt.append(d)
        for d in dfr:
            for t in dfr:
                if t[2] == d[0]:
                    dfrt.remove(d)
                    break
        for d in dfrt:
            for t in np:
                if t[1][1] <= d[1][1] and d[1][3] <= t[1][3]:
                    d[2] = t[0]
        return dfrt

    def time_pairs_ed(self, sh, ms, mr, fw):
        dfr = []
        for d in range(ms[0], ms[2] + 1):
            for t in range(ms[1], ms[3] + 1):
                if not self.sell_empty(sh.cell(row=d, column=t)):
                    dfrt = []
                    for k in range(fw[0], fw[1] + 1):
                        if not self.sell_empty(sh.cell(row=d, column=k)):
                            dfrt.append(self.lishnie_probeli(sh.cell(row=d, column=k).value, " "))
                    dfrt.extend([self.lishnie_probeli(sh.cell(row=d, column=t).value, " "),
                                 self.cell_merged_diapasone(mr, d, t)])
                    dfr.append(dfrt)
        return dfr

    def time_pairs(self, sh, ms, mr, fw, gs):
        dfr = []
        for d in range(0, min(len(self.days), len(self.dayss))):
            tr = self.koordinates(sh, self.days[d], " ", gs)
            if tr[0] == -1:
                tr = self.koordinates(sh, self.dayss[d], " ", gs)
            if tr[0] != -1:
                cr = self.cell_merged_diapasone(mr, tr[0], tr[1])
                dfr.append([self.days[d], self.time_pairs_ed(sh, [cr[0], ms[1], cr[2], ms[3]], mr, fw)])
        return dfr

    def parse_para_day(self, sh, ms, mr):
        dfr = []
        for d in range(ms[0], ms[2] + 1):
            for t in range(ms[1], ms[3] + 1):
                y = self.cell_merged_diapasone(mr, d, t)
                if not self.sell_empty(sh.cell(row=y[0], column=y[1])):
                    if not (([sh.cell(row=y[0], column=y[1]).value, 'z'] in dfr) or
                            ([sh.cell(row=y[0], column=y[1]).value, 'c'] in dfr) or
                            ([sh.cell(row=y[0], column=y[1]).value, 'cz'] in dfr)):
                        if y[0] != y[2]:
                            dfr.append([sh.cell(row=y[0], column=y[1]).value, 'cz'])
                        else:
                            if y[0] == ms[2]:
                                dfr.append([sh.cell(row=y[0], column=y[1]).value, 'z'])
                            else:
                                dfr.append([sh.cell(row=y[0], column=y[1]).value, 'c'])
        return dfr

    def parse_para_group(self, sh, mr, tpr, grt):
        dfr = []
        for d in grt:
            dfrk = [d[0]]
            for t in tpr:
                dfrt = [t[0]]
                for k in t[1]:
                    dfrt.append([k[0], self.parse_para_day(sh, [k[2][0], d[1][1], k[2][2], d[1][3]], mr)])
                dfrk.append(dfrt)
            dfr.append(dfrk)
        return dfr

    def loadbook(self, fd, cr):
        msd = self.fact_max_size(fd)
        mrt = fd.merged_cells.ranges
        gr = self.groups(fd, msd, mrt)
        if not gr:
            return
        rts = self.koordinates(fd, "День недели", "", msd)
        rts = self.cell_merged_diapasone(mrt, rts[0], rts[1])
        rtk = self.koordinates(fd, "Время занятий", " ", msd)
        crt = self.koordinates(fd, "Пары", " ", [msd[0], msd[1], rtk[0], rtk[1]])
        crt = self.cell_merged_diapasone(mrt, crt[0], crt[1])
        msd[0] = rtk[0] + 1
        msd[1] = rtk[1] + 1
        tp = self.time_pairs(fd, [msd[0], rtk[1], msd[2], rtk[1]], mrt,
                             [crt[1], crt[3]], [msd[0], rts[1], msd[2], rts[3]])
        ppg = self.parse_para_group(fd, mrt, tp, gr)
        qw = DatabaseRaspisanie(self.days, self.dayss)
        for d in tp:
            for t in d[1]:
                if qw.select_where_count('*', 'PairsTime', 'id', [t[0]]) == (0,):
                    qw.insert(qw.insert_into_time, [t[0], t[1]])
        for d in gr:
            if qw.select_where_count('*', 'GroupsOnN', ['gr', 'np', 'crs'], [d[0], d[2], cr]) == (0,):
                qw.insert(qw.insert_into_groups, [d[0], d[2], cr])
        for d in ppg:
            for t in range(1, len(d)):
                for k in range(1, len(d[t])):
                    if d[t][k]:
                        for u in range(1, len(d[t][k])):
                            for w in d[t][k][u]:
                                if qw.select_where_count('*', 'Raspisanie', ['id_pari', 'day', 'para', 'cz', 'gr'],
                                                         [d[t][k][0], self.ltr(d[t][0]), w[0], w[1], d[0]]) == (0,):
                                    qw.insert(qw.insert_into_pari, [d[t][k][0], self.ltr(d[t][0]), w[0], w[1], d[0]])
        qw.close()

    def loadfile(self, filename, crs):
        df = ox.load_workbook(filename)
        for d in df.sheetnames:
            self.loadbook(df[d], crs)
        df.close()


class ParseRaspisanie(ExcelFunctionsDefault):
    def __init__(self, jk):
        super().__init__()
        for d in jk:
            self.loadfile(d[0], d[1])

    @staticmethod
    def select(table, yslovies, values):
        qw = DatabaseRaspisanieRead()
        if not yslovies:
            er = qw.select_all(table)
            qw.close()
            return er
        er = qw.select_all_where(table, yslovies, values)
        qw.close()
        return er


class ParseRaspisanieDefault(ParseRaspisanie):
    def __init__(self):
        df = []
        for d in range(1, 5):
            df.append(['D:\\filesList\\Raspisanie_' + str(d) + '_kurs.xlsx', str(d)])
        super().__init__(df)


class Window:
    def __init__(self):
        self.w = tk.Tk()
        self.w.title("Не знаю, как назвать")
        self.w.geometry("400x510")
        self.labels = []
        self.buttons = []
        self.entries = []
        self.gt = tk.StringVar(value="Время")

    def place_label(self, side, tx, cl):
        label = tk.Label(self.w, text=tx, bg=cl)
        label.pack(side=side, fill=tk.X, padx=5, pady=2)
        self.labels.append(label)

    def printtext(self, tx, d):
        self.labels[d].config(text=tx)

    def place_button(self, side, tx, cl, fn, *n):
        button = tk.Button(self.w, text=tx, bg=cl)
        self.buttons.append(button)
        button.pack(side=side, fill=tk.X, padx=5, pady=2)
        button.config(command=lambda: fn(*n))

    def place_entry(self, side, cl):
        entry = tk.Entry(self.w, bg=cl)
        self.entries.append(entry)
        entry.pack(side=side, fill=tk.X, padx=5, pady=2)

    def place_text(self, side, cl):
        entry = tk.Text(self.w, width=25, height=5, bg=cl)
        self.entries.append(entry)
        entry.pack(side=side, fill=tk.X, padx=5, pady=2)


class WindowLoadBd(Window):
    def __init__(self):
        super().__init__()
        self.tr = ParseRaspisanie([])
        self.w.title("Редактирование базы Расписание")
        self.place_button(tk.TOP, "Загрузить дефолтный сет", "#a98307", self.button_even_load_def)
        self.place_label(tk.TOP, "Путь к папке, где лежат файлы xlsx [для всех полей ниже]", "#eb4c42")
        self.place_entry(tk.TOP, "#7b917b")
        self.entries[0].insert(1, "D:\\filesList")
        self.place_label(tk.TOP, "Напишите сюда имена файлов без формата (каждое с новой строки)", "#f75394")
        self.place_text(tk.TOP, "#7b917b")
        for d in range(1, 5):
            self.entries[1].insert(str(d) + ".0", "Raspisanie_" + str(d) + "_kurs\n")
        self.place_label(tk.TOP, "Напишите сюда имена параметров (курс), каждое с новой стркои", "#f75394")
        self.place_text(tk.TOP, "#7b917b")
        for d in range(1, 5):
            self.entries[2].insert(str(d) + ".0", str(d) + "\n")
        self.place_button(tk.TOP, "Загрузить набранный сет в таблицу Расписание", "#a98307", self.button_even_load)
        self.place_label(tk.TOP, "Имя файла для загрузки дат [также без формата]", "#f75394")
        self.place_entry(tk.TOP, "#7b917b")
        self.entries[3].insert(1, "date")
        self.place_button(tk.TOP, "Загрузить этот файл в таблицу дат", "#a98307", self.button_load_date)
        self.place_button(tk.TOP, "Сбросить базу", "#631349", self.button_clear_base)
        self.place_button(tk.TOP, "Таблица для логов: Raspisanie", "#497e76", self.button_edit_logs)
        self.place_button(tk.TOP, "Логи", "#9aceeb", self.button_print_logs)
        self.tablesdorlogs = ['Raspisanie', 'ChislitelZnamenatelWeeks', 'PairsTime', 'GroupsOnN', 'DayOfWeek']
        self.position = 0
        self.w.mainloop()

    def button_even_load_def(self):
        self.tr = ParseRaspisanieDefault()

    def button_even_load(self):
        jk = []
        pth = self.entries[0].get()
        sp = self.entries[1].get("1.0", "end-1c").split("\n")
        ht = self.entries[2].get("1.0", "end-1c").split("\n")
        uy = []
        for d in range(0, min(len(sp), len(ht))):
            if sp[d] != "":
                if not sp[d].endswith(".xlsx"):
                    sp[d] = sp[d] + ".xlsx"
                uy.append([pth + "\\" + sp[d], ht[d]])
        sp = []
        for d in uy:
            if os.path.exists(d[0]):
                jk.append([d[0], d[1]])
            else:
                sp.append(d)
        ou = "Не найден путь к следующим файлам: \n"
        for d in sp:
            ou = ou + d + "\n"
        if jk:
            self.tr = ParseRaspisanie(jk)
        else:
            ou = "Все файлы не найдены.\n Убедитесь, что установлена та папка.\n"
        if sp:
            mb.showwarning("Файлы не найдены", ou)

    def button_load_date(self):
        pth = self.entries[0].get()
        fl = self.entries[3].get()
        if not fl.endswith(".xlsx"):
            fl = fl + ".xlsx"
        try:
            df = ox.load_workbook(pth + "\\" + fl)
            vls = []
            for d in df.sheetnames:
                fd = df[d]
                fs = self.tr.fact_max_size(fd)
                for t in range(fs[0], fs[2] + 1):
                    ytki = True
                    for k in range(fs[1], fs[3]):
                        for u in range(fs[1] + 1, fs[3] + 1):
                            if not (self.tr.sell_empty(fd.cell(row=t, column=k)) or
                                    self.tr.sell_empty(fd.cell(row=t, column=u))):
                                vls.append([fd.cell(row=t, column=k).value, fd.cell(row=t, column=u).value])
                                ytki = False
                                break
                        if not ytki:
                            break
            df.close()
            for d in vls:
                if d[1] == 'числитель':
                    d[1] = 'c'
                else:
                    d[1] = 'z'
            qw = DatabaseRaspisanieRead()
            for d in vls:
                if qw.select_where_count('*', 'ChislitelZnamenatelWeeks', ['week', 'cz'], d) == (0,):
                    qw.insert(qw.insert_into_weeks, d)
            qw.close()
        except Exception:
            mb.showwarning("Данного файла не существует", "Внимательно проверьте путь к файлу ещё раз")

    @staticmethod
    def button_clear_base():
        if os.path.exists("Raspisanie.db"):
            os.remove("Raspisanie.db")

    def button_print_logs(self):
        er = self.tr.select(self.tablesdorlogs[self.position], [], [])
        iu = "<table border='1'>"
        for d in er:
            print(d)
            iu = iu + "<tr>"
            for t in d:
                iu = iu + "<td>" + str(t) + "</td>"
            iu = iu + "</tr>"
        f = open('try.txt', 'w')
        f.write(iu + "</table>")
        f.close()

    def button_edit_logs(self):
        self.position = self.position + 1
        if self.position >= len(self.tablesdorlogs):
            self.position = 0
        self.buttons[4].config(text="Таблица для логов: " + self.tablesdorlogs[self.position])


def edit_fg(fh):
    f = open('try.txt', 'w')
    f.write(fh)
    f.close()


def echer():
    guy = http.client.HTTPConnection("127.0.0.1:8000")
    guy.request("GET", "/")
    response = guy.getresponse()
    if response.status == 200:
        data = response.read().decode("utf-8")
        guy.close()
        return data
    guy.close()
    return ""


def loght():
    sbp.run(["python", "servakget.py"])
    edit_fg("Порубаем в капусту")


def server():
    sbp.run(["python", "socketbud.py"])


thg = [trg.Thread(target=WindowLoadBd), trg.Thread(target=loght), trg.Thread(target=server)]
for thr in thg:
    thr.start()
